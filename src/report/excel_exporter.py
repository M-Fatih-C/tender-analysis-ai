"""
TenderAI Excel Exporter.

Analiz sonuçlarını formatlanmış Excel dosyasına export eder.
"""

import io
import json
import logging

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)

_HEADER_FILL = PatternFill(start_color="667eea", end_color="764ba2", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_RISK_FILLS = {
    "KRİTİK": PatternFill(start_color="c0392b", fill_type="solid"),
    "YÜKSEK": PatternFill(start_color="e74c3c", fill_type="solid"),
    "ORTA": PatternFill(start_color="f39c12", fill_type="solid"),
    "DÜŞÜK": PatternFill(start_color="27ae60", fill_type="solid"),
}
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _safe(data, key, default="—"):
    if isinstance(data, dict):
        return data.get(key, default)
    return default


def _add_header(ws, row, cols):
    for col_idx, text in enumerate(cols, 1):
        cell = ws.cell(row=row, column=col_idx, value=text)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = _THIN_BORDER


class ExcelExporter:
    """Analiz sonuçlarını Excel'e export eder."""

    def export(self, result: dict, file_name: str = "Analiz") -> bytes:
        """Detaylı Excel raporu oluştur."""
        if isinstance(result, str):
            try:
                result = json.loads(result)
            except Exception:
                result = {}

        wb = Workbook()

        # Sheet 1: Yönetici Özeti
        ws = wb.active
        ws.title = "Yönetici Özeti"
        self._sheet_summary(ws, result, file_name)

        # Sheet 2: Risk Analizi
        ws2 = wb.create_sheet("Risk Analizi")
        self._sheet_risks(ws2, result)

        # Sheet 3: Gerekli Belgeler
        ws3 = wb.create_sheet("Gerekli Belgeler")
        self._sheet_documents(ws3, result)

        # Sheet 4: Ceza Maddeleri
        ws4 = wb.create_sheet("Ceza Maddeleri")
        self._sheet_penalties(ws4, result)

        # Sheet 5: Mali Özet
        ws5 = wb.create_sheet("Mali Özet")
        self._sheet_financial(ws5, result)

        # Sheet 6: Süre Analizi
        ws6 = wb.create_sheet("Süre Analizi")
        self._sheet_timeline(ws6, result)

        # Auto-fit columns
        for ws in wb.worksheets:
            for col in ws.columns:
                max_len = 0
                for cell in col:
                    try:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    except Exception:
                        pass
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _sheet_summary(self, ws, result, file_name):
        ws.cell(row=1, column=1, value="TenderAI Analiz Raporu").font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=f"Dosya: {file_name}")
        ws.cell(row=3, column=1, value=f"Risk Skoru: {result.get('risk_score', '—')}")
        ws.cell(row=4, column=1, value=f"Risk Seviyesi: {result.get('risk_level', '—')}")

        exec_data = _safe(result, "executive_summary", {})
        ws.cell(row=6, column=1, value="Özet:").font = Font(bold=True)
        ws.cell(row=7, column=1, value=_safe(exec_data, "ozet", ""))

        guclu = _safe(exec_data, "guclu_yanlar", [])
        if guclu:
            ws.cell(row=9, column=1, value="Güçlü Yanlar:").font = Font(bold=True)
            for i, item in enumerate(guclu):
                ws.cell(row=10 + i, column=1, value=f"✅ {item}")

    def _sheet_risks(self, ws, result):
        risk_data = _safe(result, "risk_analysis", {})
        riskler = _safe(risk_data, "riskler", [])
        _add_header(ws, 1, ["Kategori", "Başlık", "Açıklama", "Seviye", "Madde", "Öneri"])
        for i, r in enumerate(riskler, 2):
            if not isinstance(r, dict):
                continue
            ws.cell(row=i, column=1, value=r.get("kategori", ""))
            ws.cell(row=i, column=2, value=r.get("baslik", ""))
            ws.cell(row=i, column=3, value=r.get("aciklama", ""))
            sev = r.get("seviye", "ORTA")
            cell = ws.cell(row=i, column=4, value=sev)
            if sev in _RISK_FILLS:
                cell.fill = _RISK_FILLS[sev]
                cell.font = Font(color="FFFFFF", bold=True)
            ws.cell(row=i, column=5, value=r.get("madde_referans", ""))
            ws.cell(row=i, column=6, value=r.get("oneri", ""))

    def _sheet_documents(self, ws, result):
        docs = _safe(result, "required_documents", {})
        zorunlu = _safe(docs, "zorunlu_belgeler", [])
        _add_header(ws, 1, ["Belge Adı", "Açıklama", "Kategori", "Nereden Alınır", "Süre"])
        for i, d in enumerate(zorunlu, 2):
            if not isinstance(d, dict):
                continue
            ws.cell(row=i, column=1, value=d.get("belge_adi", ""))
            ws.cell(row=i, column=2, value=d.get("aciklama", ""))
            ws.cell(row=i, column=3, value=d.get("kategori", ""))
            ws.cell(row=i, column=4, value=d.get("nereden_alinir", ""))
            ws.cell(row=i, column=5, value=d.get("tahmini_sure", ""))

    def _sheet_penalties(self, ws, result):
        pen = _safe(result, "penalty_clauses", {})
        cezalar = _safe(pen, "cezalar", [])
        _add_header(ws, 1, ["Madde", "Tür", "Oran/Tutar", "Açıklama", "Seviye", "Senaryo", "Öneri"])
        for i, c in enumerate(cezalar, 2):
            if not isinstance(c, dict):
                continue
            ws.cell(row=i, column=1, value=c.get("madde_no", ""))
            ws.cell(row=i, column=2, value=c.get("ceza_turu", ""))
            ws.cell(row=i, column=3, value=c.get("miktar_oran", ""))
            ws.cell(row=i, column=4, value=c.get("aciklama", ""))
            sev = c.get("risk_seviyesi", "ORTA")
            cell = ws.cell(row=i, column=5, value=sev)
            if sev in _RISK_FILLS:
                cell.fill = _RISK_FILLS[sev]
                cell.font = Font(color="FFFFFF", bold=True)
            ws.cell(row=i, column=6, value=c.get("senaryo", ""))
            ws.cell(row=i, column=7, value=c.get("oneri", ""))

    def _sheet_financial(self, ws, result):
        fin = _safe(result, "financial_summary", {})
        ws.cell(row=1, column=1, value="Mali Özet").font = Font(bold=True, size=13)
        items = [
            ("Tahmini İhale Bedeli", _safe(fin, "tahmini_ihale_bedeli")),
            ("Geçici Teminat", _safe(fin, "gecici_teminat")),
            ("Kesin Teminat", _safe(fin, "kesin_teminat")),
            ("Ödeme Koşulları", _safe(fin, "odeme_kosullari")),
        ]
        for i, (label, val) in enumerate(items, 3):
            ws.cell(row=i, column=1, value=label).font = Font(bold=True)
            ws.cell(row=i, column=2, value=str(val))

    def _sheet_timeline(self, ws, result):
        tl = _safe(result, "timeline_analysis", {})
        ws.cell(row=1, column=1, value="Süre Analizi").font = Font(bold=True, size=13)
        ws.cell(row=3, column=1, value="Toplam Süre").font = Font(bold=True)
        ws.cell(row=3, column=2, value=_safe(tl, "toplam_is_suresi"))

        milestones = _safe(tl, "milestones", [])
        if milestones:
            _add_header(ws, 5, ["Aşama", "Süre", "Tamamlanma"])
            for i, ms in enumerate(milestones, 6):
                if isinstance(ms, dict):
                    ws.cell(row=i, column=1, value=ms.get("asama", ""))
                    ws.cell(row=i, column=2, value=ms.get("sure", ""))
                    ws.cell(row=i, column=3, value=ms.get("tamamlanma", ""))

    def export_comparison(self, comparison_result: dict) -> bytes:
        """Karşılaştırma tablosunu Excel'e export et."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Karşılaştırma"

        rows = comparison_result.get("rows", [])
        if not rows:
            ws.cell(row=1, column=1, value="Karşılaştırma verisi yok")
            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue()

        headers = ["İhale", "Risk Skoru", "Bedel", "Teminat", "Süre", "Belge", "Ceza", "Tavsiye"]
        _add_header(ws, 1, headers)

        for i, r in enumerate(rows, 2):
            ws.cell(row=i, column=1, value=r.get("name", ""))
            ws.cell(row=i, column=2, value=r.get("risk_score", 0))
            ws.cell(row=i, column=3, value=r.get("bedel", ""))
            ws.cell(row=i, column=4, value=r.get("teminat", ""))
            ws.cell(row=i, column=5, value=r.get("sure", ""))
            ws.cell(row=i, column=6, value=r.get("belge_sayisi", 0))
            ws.cell(row=i, column=7, value=r.get("ceza_sayisi", 0))
            ws.cell(row=i, column=8, value=r.get("tavsiye", ""))

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
